import os
import re
import logging
import unicodedata
from urllib.parse import unquote

# Load environment variables from .env file if it exists
# This should be done before any other imports that might use env vars
from dotenv import load_dotenv
load_dotenv()

from flask import Flask, render_template, request, jsonify, session, redirect, url_for, send_file
import pandas as pd
from .utils import (calculate_standings_by_division, get_highest_scoring_games, 
                   load_game_data, get_top_players_by_score, get_team_performance_stats,
                   get_top_scorers, get_highest_single_game_score, get_top_three_pointers, 
                   get_top_foulers, get_referee_statistics, get_referee_fouls_per_game,
                   get_referees_least_fouls_per_game, get_biggest_wins, get_biggest_leads,
                   get_most_tie_scores, get_most_lead_changes, get_longest_duration_games,
                   get_player_shooting_efficiency, get_biggest_scoring_streaks,
                   get_starting_five_vs_bench_stats, get_double_digit_scorers, get_consistent_scorers,
                   get_player_game_impact_analysis, get_player_foul_impact_analysis,
                   get_best_player_combinations, get_referee_game_impact_analysis, get_all_fixtures_data,
                   get_fixtures_matrix_data, get_data_source_info, get_season_info, 
                   get_website_config, list_available_archives, import_season_archive, export_season_archive,
                   get_all_players_list, get_player_detail_stats, get_game_details, generate_game_review, get_referee_detail_stats,
                   get_team_detail_stats, get_all_referees_list, get_all_games_list,
                   get_player_hover_stats, get_team_hover_stats, get_referee_hover_stats, get_game_hover_stats,
                   get_division_hover_stats, calculate_referee_performance_index, get_closest_games_by_team,
                   extract_age_sex_group_from_division, get_team_name_with_group_suffix, CSV_FILEPATH,
                   get_team_three_pointers_stats, get_team_fouls_stats, get_team_fouls_trend_data,
                   get_team_highest_single_game_scores,
                   get_team_free_throw_stats, get_team_double_digit_scorers_stats, get_team_consistency_stats,
                   get_team_starting_vs_bench_stats)
from .version import get_version_info
from .user_database import (authenticate_user, get_user_preferences, update_user_preferences,
                            create_user, list_users, update_user_password, delete_user, update_user_level,
                            get_users_with_login_info, get_recent_login_logs, get_login_statistics,
                            get_foul_weights, update_foul_weights)

app = Flask(__name__, template_folder='../templates', static_folder='../logos', static_url_path='/logos')

# Valid theme options for the application
VALID_THEMES = ['default', 'ocean', 'sunset', 'forest', 'minimal', 'cherry']

# Rate limiting constants
FOUL_WEIGHTS_UPDATE_RATE_LIMIT = '30 per hour'

# Configure logging for tracking code validation
logger = logging.getLogger(__name__)

# Configure rate limiting for brute-force protection
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address

# Initialize rate limiter
limiter = Limiter(
    app=app,
    key_func=get_remote_address,
    default_limits=["200 per day", "50 per hour"],
    storage_uri="memory://",
    strategy="fixed-window"
)

# Custom error handler for rate limit exceeded
@app.errorhandler(429)
def ratelimit_handler(e):
    """Handle rate limit exceeded errors with a user-friendly message"""
    # Check if this is a login endpoint
    if '/login' in request.path:
        # Return to login page with error message
        error_msg = "Too many login attempts. Please try again in 15 minutes."
        return render_template('login.html', error=error_msg), 429
    
    # For other endpoints, return a generic error
    return jsonify({
        'error': 'Rate limit exceeded',
        'message': 'Too many requests. Please try again later.'
    }), 429


def validate_tracking_code(code):
    """
    Validate tracking code from environment variable for basic security.
    
    This function performs basic validation to ensure the tracking code:
    - Is not excessively long (prevents DoS)
    - Contains properly formatted script tags
    - Doesn't contain obvious malicious patterns
    
    Note: This is not a comprehensive XSS prevention mechanism. The tracking code
    should only be set from trusted sources (e.g., MyDevil panel). Never allow
    user input to set this value.
    
    Args:
        code (str): The tracking code to validate
        
    Returns:
        str: The validated code, or empty string if invalid
    """
    if not code or not isinstance(code, str):
        return ''
    
    # Check length (tracking codes shouldn't be huge)
    if len(code) > 10000:  # 10KB max
        logger.warning("MYDEVIL_STATS_CODE exceeds maximum length (10KB), ignoring")
        return ''
    
    # Validate script tag format - should have both opening and closing tags
    if '<script' not in code.lower():
        logger.warning("MYDEVIL_STATS_CODE doesn't contain <script> tag, ignoring")
        return ''
    
    if '</script>' not in code.lower():
        logger.warning("MYDEVIL_STATS_CODE doesn't contain closing </script> tag, ignoring")
        return ''
    
    # Check for dangerous patterns that could indicate XSS attempts
    # Note: This is a basic blocklist, not a comprehensive security mechanism
    dangerous_patterns = [
        'javascript:',      # javascript: protocol
        'data:',            # data: protocol
        'vbscript:',        # vbscript: protocol
        'onerror=',         # event handlers
        'onload=',
        'onclick=',
        'onmouseover=',
        'onfocus=',
        'onblur=',
        'onchange=',
        'onsubmit=',
        '<iframe',          # iframe injection
        'document.cookie',  # cookie theft
        'eval(',            # code execution
        'expression(',      # IE CSS expressions
    ]
    code_lower = code.lower()
    for pattern in dangerous_patterns:
        if pattern in code_lower:
            logger.warning(
                f"MYDEVIL_STATS_CODE contains potentially dangerous pattern '{pattern}', ignoring"
            )
            return ''
    
    return code

# Set secret key for session management
# In production, SECRET_KEY should be set via environment variable
# For development, generate a random key if not set
if not os.environ.get('SECRET_KEY'):
    import secrets
    app.secret_key = secrets.token_hex(32)
else:
    app.secret_key = os.environ.get('SECRET_KEY')

# Configure session to persist properly across page navigation
from datetime import timedelta
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=31)  # Session expires after 31 days
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'  # Allow cookies on same-site navigation
app.config['SESSION_COOKIE_HTTPONLY'] = True  # Prevent JavaScript access for security
# Enable secure cookies in production when HTTPS is forced
app.config['SESSION_COOKIE_SECURE'] = os.environ.get('FORCE_HTTPS', 'false').lower() == 'true'
app.config['SESSION_REFRESH_EACH_REQUEST'] = True  # Refresh session on each request to keep it alive

# Admin and User authentication
from functools import wraps

def is_admin_authenticated():
    """Check if the current user is authenticated as admin"""
    return session.get('user_level') == 'admin'

def is_user_authenticated():
    """Check if the current user is authenticated as regular user or admin"""
    user_level = session.get('user_level', 'guest')
    return user_level in ('user', 'admin')

def get_user_level():
    """Get the current user's authorization level: 'guest', 'user', or 'admin'"""
    return session.get('user_level', 'guest')

def login_required(f):
    """Decorator to require admin authentication for a route"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not is_admin_authenticated():
            return jsonify({'success': False, 'error': 'Admin authentication required'}), 401
        return f(*args, **kwargs)
    return decorated_function

def user_required(f):
    """Decorator to require user or admin authentication for a route"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not is_user_authenticated():
            return redirect(url_for('login', next=request.url))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    """Decorator to require admin authentication for a route"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not is_admin_authenticated():
            return redirect(url_for('login', next=request.url))
        return f(*args, **kwargs)
    return decorated_function

# Context processor to make season info available to all templates
@app.context_processor
def inject_season_info():
    """Make season information available to all templates"""
    season_info = get_season_info()
    website_config = get_website_config()
    version_info = get_version_info()
    
    # Get user preferences from session
    user_prefs = {
        'division': session.get('preferred_division'),
        'team': session.get('preferred_team'),
        'player': session.get('preferred_player'),
        'theme': session.get('preferred_theme', 'default')
    }
    
    # Get MyDevil statistics tracking code from environment variable and validate it
    mydevil_stats_code = validate_tracking_code(os.environ.get('MYDEVIL_STATS_CODE', ''))
    
    return {
        'season_info': season_info,
        'website_config': website_config,
        'version_info': version_info,
        'user_prefs': user_prefs,
        'mydevil_stats_code': mydevil_stats_code,
        'is_admin_authenticated': is_admin_authenticated(),
        'is_user_authenticated': is_user_authenticated(),
        'user_level': get_user_level(),
        'current_username': session.get('username')
    }

# Logo utility functions
def normalize_team_name(team_name):
    """Normalize team name for file naming
    
    Converts accented characters to their base form (é -> e, ä -> a, etc.)
    before removing remaining special characters and converting to lowercase.
    """
    if not team_name:
        return ""
    
    # First, normalize accents to their base characters
    # NFD = Canonical Decomposition (separates base character from accent)
    # Filter out combining marks (category 'Mn') to remove accents
    normalized = ''.join(
        c for c in unicodedata.normalize('NFD', str(team_name))
        if unicodedata.category(c) != 'Mn'
    )
    
    # Then remove any remaining special characters
    normalized = re.sub(r'[^a-zA-Z0-9\s]', '', normalized)
    normalized = normalized.lower().replace(' ', '-')
    return normalized

def get_team_logo_url(team_name):
    """Get the logo URL for a given team name for use in templates"""
    if not team_name:
        return None
        
    normalized_name = normalize_team_name(team_name)
    logos_dir = "logos"
    
    # Check for PNG files (our standard format) - check both .png and .PNG
    for png_ext in ['.png', '.PNG']:
        png_file = f"{normalized_name}{png_ext}"
        png_path = os.path.join(logos_dir, png_file)
        if os.path.exists(png_path):
            return f"/logos/{png_file}"
    
    # Check for other formats
    for ext in ['.jpg', '.jpeg', '.gif', '.svg']:
        logo_file = f"{normalized_name}{ext}"
        logo_path = os.path.join(logos_dir, logo_file)
        if os.path.exists(logo_path):
            return f"/logos/{logo_file}"
    
    # Special case for Racing teams - check RAC.jpg
    if 'racing' in normalized_name.lower():
        rac_path = os.path.join(logos_dir, "RAC.jpg")
        if os.path.exists(rac_path):
            return "/logos/RAC.jpg"
    
    return None

# Make logo function available to templates
app.jinja_env.globals.update(get_team_logo_url=get_team_logo_url)

# Make age/sex group functions available to templates
app.jinja_env.globals.update(
    extract_age_sex_group_from_division=extract_age_sex_group_from_division,
    get_team_name_with_group_suffix=get_team_name_with_group_suffix
)

# Load and process the data
data = load_game_data()
data_source_info = get_data_source_info()

if not data.empty:
    # Clean the data - remove unnamed column if it exists
    if 'Unnamed: 0' in data.columns:
        data = data.drop('Unnamed: 0', axis=1)
    # Extract unique divisions
    divisions = data['GameDivisionDisplay'].unique()
    # Extract unique teams from player stats
    from src.utils import extract_all_player_stats
    player_stats = extract_all_player_stats(data)
    if not player_stats.empty:
        teams = sorted(player_stats['Team'].unique())
    else:
        teams = []
else:
    divisions = []
    teams = []

def filter_data_by_division(data, selected_division):
    """
    Filter data by division, with special handling for DIVISIONS_1_4.
    
    Parameters:
    data (DataFrame): The game data
    selected_division (str): The division filter (can be None, empty string, specific division, or "DIVISIONS_1_4")
    
    Returns:
    DataFrame: Filtered data
    
    Behavior:
    - None or empty string: returns all data (no filter)
    - "DIVISIONS_1_4": returns only games from divisions "M-Division 1:", "M-Division 2:", "M-Division 3:", and "M-Division 4:"
    - Specific division name: returns games from that division only
    """
    if not selected_division:
        return data
    
    if selected_division == "DIVISIONS_1_4":
        # Filter to only M-Division 1:, 2:, 3:, 4:
        return data[data['GameDivisionDisplay'].isin([
            'M-Division 1:', 'M-Division 2:', 'M-Division 3:', 'M-Division 4:'
        ])]
    else:
        # Regular single division filter
        return data[data['GameDivisionDisplay'] == selected_division]

@app.route('/', methods=['GET', 'POST'])
def index():
    """Home page with navigation tiles"""
    return render_template('index.html', 
                         divisions=divisions,
                         data_source_info=data_source_info)

@app.route('/standings', methods=['GET', 'POST'])
def standings():
    """Division standings page"""
    # Get division from form, query params, or user preferences
    selected_division = request.form.get('division') or request.args.get('division') or session.get('preferred_division')
    standings = None

    if selected_division and not data.empty:
        # Calculate standings for the selected division
        standings = calculate_standings_by_division(data, selected_division)

    return render_template('standings.html', 
                         divisions=divisions, 
                         standings=standings,
                         selected_division=selected_division,
                         data_source_info=data_source_info)

@app.route('/statistics', methods=['GET', 'POST'])
@user_required
def statistics():
    """Game Statistics page with division filtering"""
    if data.empty:
        return render_template('statistics.html', 
                             error="No data available",
                             divisions=divisions,
                             data_source_info=data_source_info)
    
    # Get selected division from form, query parameter, or user preferences
    selected_division = request.form.get('division') or request.args.get('division') or session.get('preferred_division')
    
    # Filter data based on division selection
    filtered_data = filter_data_by_division(data, selected_division)
    
    # Get game statistics with filtered data
    highest_games = get_highest_scoring_games(filtered_data, 10, division=None)
    biggest_wins = get_biggest_wins(filtered_data, 10, division=None)
    biggest_leads = get_biggest_leads(filtered_data, 10, division=None)
    biggest_streaks = get_biggest_scoring_streaks(filtered_data, 10, division=None)
    most_ties = get_most_tie_scores(filtered_data, 10, division=None)
    most_lead_changes = get_most_lead_changes(filtered_data, 10, division=None)
    longest_duration_games = get_longest_duration_games(filtered_data, 20, division=None)
    
    return render_template('statistics.html', 
                         highest_games=highest_games,
                         biggest_wins=biggest_wins,
                         biggest_leads=biggest_leads,
                         biggest_streaks=biggest_streaks,
                         most_ties=most_ties,
                         most_lead_changes=most_lead_changes,
                         longest_duration_games=longest_duration_games,
                         divisions=divisions,
                         selected_division=selected_division,
                         data_source_info=data_source_info)

@app.route('/team-stats', methods=['GET', 'POST'])
@user_required
def team_stats():
    """Complete team statistics page with division filter"""
    if data.empty:
        return render_template('team_stats.html', error="No data available", data_source_info=data_source_info)
    
    # Get selected division from form or user preferences
    selected_division = request.form.get('division') or session.get('preferred_division')
    
    # Filter data based on division selection
    filtered_data = filter_data_by_division(data, selected_division)
    
    # Get team performance stats from filtered data
    team_performance = get_team_performance_stats(filtered_data)
    
    # Get additional team statistics for tabs
    team_three_pointers = get_team_three_pointers_stats(filtered_data, 20)
    team_fouls = get_team_fouls_stats(filtered_data, 20)
    team_highest_scores = get_team_highest_single_game_scores(filtered_data, 20)
    team_free_throws = get_team_free_throw_stats(filtered_data, 20)
    team_double_digit = get_team_double_digit_scorers_stats(filtered_data, 10)
    team_consistency = get_team_consistency_stats(filtered_data, 5)
    team_starter_bench = get_team_starting_vs_bench_stats(filtered_data)
    
    return render_template('team_stats.html', 
                         team_stats=team_performance,
                         team_three_pointers=team_three_pointers,
                         team_fouls=team_fouls,
                         team_highest_scores=team_highest_scores,
                         team_free_throws=team_free_throws,
                         team_double_digit=team_double_digit,
                         team_consistency=team_consistency,
                         team_starter_bench=team_starter_bench,
                         divisions=divisions,
                         selected_division=selected_division,
                         data_source_info=data_source_info)

@app.route('/team-stats/export-fouls')
@user_required
def export_team_fouls():
    """Export team fouls statistics to Excel file"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from flask import send_file
        import io
        
        # Get selected division from query parameter or user preferences
        selected_division = request.args.get('division') or session.get('preferred_division')
        
        # Filter data based on division selection
        filtered_data = filter_data_by_division(data, selected_division)
        
        # Get team fouls stats (all teams for comprehensive export)
        team_fouls = get_team_fouls_stats(filtered_data, top_n=500)  # Max 500 teams should cover all leagues
        
        if team_fouls.empty:
            return "No data available", 404
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Team Fouls Statistics"
        
        # Header info
        ws['A1'] = "Team Fouls Statistics"
        ws['A1'].font = Font(bold=True, size=16)
        
        division_text = f"Division: {selected_division}" if selected_division else "All Divisions"
        ws['A2'] = division_text
        ws['A2'].font = Font(bold=True, size=12)
        
        # Header row (starting at row 4)
        row = 4
        headers = [
            'Rank', 'Team', 'Total Fouls', 'Fouls per Game', 
            'P', 'P1', 'P2', 'P3', 'T1', 'U1', 'U2', 'U3', 'GD',
            'Weighted Total', 'Games Played', 'Total Points'
        ]
        
        # Style header row (color matches app theme gradient)
        HEADER_COLOR = "667eea"  # Primary theme color
        header_fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Data rows
        for idx, (_, team_row) in enumerate(team_fouls.iterrows(), start=1):
            row += 1
            values = [
                idx,  # Rank
                team_row['Team'],
                int(team_row['TotalFouls']),
                round(team_row['AvgFoulsPerGame'], 1),
                int(team_row['PFouls']),
                int(team_row['P1Fouls']),
                int(team_row['P2Fouls']),
                int(team_row['P3Fouls']),
                int(team_row['T1Fouls']),
                int(team_row['U1Fouls']),
                int(team_row['U2Fouls']),
                int(team_row['U3Fouls']),
                int(team_row['GDFouls']),
                int(team_row['WeightedTotalFouls']),
                int(team_row['TotalGames']),
                int(team_row['TotalPoints'])
            ]
            
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                if col == 2:  # Team name
                    cell.alignment = Alignment(horizontal='left')
                else:
                    cell.alignment = Alignment(horizontal='center')
        
        # Add legend
        row += 3
        ws[f'A{row}'] = "Foul Types Legend:"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        ws[f'A{row}'] = "P=Personal, P1=Personal 1, P2=Personal 2, P3=Personal 3"
        row += 1
        ws[f'A{row}'] = "T1=Technical, U1=Unsportsmanlike 1, U2=Unsportsmanlike 2, U3=Unsportsmanlike 3"
        row += 1
        ws[f'A{row}'] = "GD=Game Disqualification"
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 8   # Rank
        ws.column_dimensions['B'].width = 30  # Team
        ws.column_dimensions['C'].width = 14  # Total Fouls
        ws.column_dimensions['D'].width = 15  # Fouls per Game
        for col in ['E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:  # Foul types
            ws.column_dimensions[col].width = 8
        ws.column_dimensions['N'].width = 15  # Weighted Total
        ws.column_dimensions['O'].width = 14  # Games
        ws.column_dimensions['P'].width = 14  # Points
        
        # Save to BytesIO
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Create filename
        division_suffix = f"_{selected_division.replace(' ', '_')}" if selected_division else "_All"
        filename = f"Team_Fouls_Statistics{division_suffix}.xlsx"
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logger.error(f"Error exporting team fouls: {str(e)}")
        return f"Error exporting data: {str(e)}", 500

@app.route('/team-stats/fouls-trend-data')
@user_required
def get_fouls_trend_data():
    """API endpoint to get fouls trend data for selected teams"""
    try:
        # Get team names from query parameter (can be comma-separated list)
        team_names_str = request.args.get('teams', '')
        selected_division = request.args.get('division') or session.get('preferred_division')
        
        # Filter data based on division selection
        filtered_data = filter_data_by_division(data, selected_division)
        
        # Parse team names
        team_names = [name.strip() for name in team_names_str.split(',') if name.strip()]
        
        # Get trend data
        trend_data = get_team_fouls_trend_data(filtered_data, team_names if team_names else None)
        
        return jsonify(trend_data)
        
    except Exception as e:
        logger.error(f"Error getting fouls trend data: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/team-stats/export-fouls-trend')
@user_required
def export_fouls_trend():
    """Export fouls trend data to Excel file"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from flask import send_file
        import io
        
        # Get team names from query parameter (can be comma-separated list)
        team_names_str = request.args.get('teams', '')
        selected_division = request.args.get('division') or session.get('preferred_division')
        
        # Filter data based on division selection
        filtered_data = filter_data_by_division(data, selected_division)
        
        # Parse team names
        team_names = [name.strip() for name in team_names_str.split(',') if name.strip()]
        
        # Get trend data
        trend_data = get_team_fouls_trend_data(filtered_data, team_names if team_names else None)
        
        if not trend_data:
            return "No data available", 404
        
        # Create workbook
        wb = Workbook()
        
        # Track sheet names to ensure uniqueness
        sheet_names_used = set()
        
        # Create a sheet for each team
        for idx, (team_name, team_data) in enumerate(trend_data.items()):
            # Handle both new and old data structure
            games = team_data.get('games', team_data) if isinstance(team_data, dict) else team_data
            statistics = team_data.get('statistics', {}) if isinstance(team_data, dict) else {}
            
            # Create unique sheet name (Excel limit: 31 chars)
            base_name = team_name[:28]  # Leave room for potential suffix
            sheet_name = base_name
            counter = 1
            while sheet_name in sheet_names_used:
                suffix = f"_{counter}"
                sheet_name = base_name[:31-len(suffix)] + suffix
                counter += 1
            sheet_names_used.add(sheet_name)
            
            if idx == 0:
                ws = wb.active
                ws.title = sheet_name
            else:
                ws = wb.create_sheet(title=sheet_name)
            
            # Header info
            ws['A1'] = f"Fouls Trend for {team_name}"
            ws['A1'].font = Font(bold=True, size=14)
            
            division_text = f"Division: {selected_division}" if selected_division else "All Divisions"
            ws['A2'] = division_text
            ws['A2'].font = Font(bold=True, size=11)
            
            # Add statistics summary if available
            row = 3
            if statistics:
                ws[f'A{row}'] = f"Statistics Summary"
                ws[f'A{row}'].font = Font(bold=True, size=11)
                row += 1
                
                stats_text = (
                    f"Average: {statistics.get('average', 'N/A')} fouls/game | "
                    f"Trend: {statistics.get('trend_direction', 'N/A')} {statistics.get('trend_indicator', '')} | "
                    f"First Half Avg: {statistics.get('first_half_avg', 'N/A')} | "
                    f"Second Half Avg: {statistics.get('second_half_avg', 'N/A')} | "
                    f"Change: {statistics.get('change_percent', 'N/A')}%"
                )
                ws[f'A{row}'] = stats_text
                ws[f'A{row}'].font = Font(size=10)
                row += 1
            
            # Column headers (starting at next row)
            row += 1
            headers = [
                'Game #', 'Date', 'Game ID', 'Total Fouls',
                'P', 'P1', 'P2', 'P3', 'T1', 'U1', 'U2', 'U3', 'GD'
            ]
            
            # Style header row
            HEADER_COLOR = "667eea"
            header_fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Data rows
            for game in games:
                row += 1
                fouls = game['fouls_by_type']
                values = [
                    game['game_number'],
                    game['date'],
                    game['game_id'],
                    game['total_fouls'],
                    fouls.get('P', 0),
                    fouls.get('P1', 0),
                    fouls.get('P2', 0),
                    fouls.get('P3', 0),
                    fouls.get('T1', 0),
                    fouls.get('U1', 0),
                    fouls.get('U2', 0),
                    fouls.get('U3', 0),
                    fouls.get('GD', 0)
                ]
                
                for col, value in enumerate(values, start=1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = border
                    if col == 2:  # Date
                        cell.alignment = Alignment(horizontal='left')
                    elif col == 3:  # Game ID
                        cell.alignment = Alignment(horizontal='left')
                    else:
                        cell.alignment = Alignment(horizontal='center')
            
            # Add legend
            row += 3
            ws[f'A{row}'] = "Foul Types Legend:"
            ws[f'A{row}'].font = Font(bold=True, size=11)
            row += 1
            ws[f'A{row}'] = "P=Personal, P1=Personal 1, P2=Personal 2, P3=Personal 3"
            row += 1
            ws[f'A{row}'] = "T1=Technical, U1=Unsportsmanlike 1, U2=Unsportsmanlike 2, U3=Unsportsmanlike 3"
            row += 1
            ws[f'A{row}'] = "GD=Game Disqualification"
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 8   # Game #
            ws.column_dimensions['B'].width = 20  # Date
            ws.column_dimensions['C'].width = 12  # Game ID
            ws.column_dimensions['D'].width = 12  # Total Fouls
            for col in ['E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
                ws.column_dimensions[col].width = 6
        
        # Save to BytesIO
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Create filename
        if team_names and len(team_names) == 1:
            team_suffix = f"_{team_names[0].replace(' ', '_')}"
        elif team_names and len(team_names) <= 3:
            team_suffix = f"_{'_'.join([t.replace(' ', '_')[:10] for t in team_names])}"
        else:
            team_suffix = "_Multiple_Teams"
        
        division_suffix = f"_{selected_division.replace(' ', '_')}" if selected_division else ""
        filename = f"Team_Fouls_Trend{team_suffix}{division_suffix}.xlsx"
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logger.error(f"Error exporting fouls trend: {str(e)}")
        return f"Error exporting data: {str(e)}", 500

@app.route('/team-detail')
@user_required
def team_detail():
    """Individual team detail page with comprehensive statistics"""
    if data.empty:
        return render_template('team_detail.html', error="No data available", data_source_info=data_source_info)
    
    # Get all unique teams
    home_teams = set(data['HomeTeamName'].unique())
    away_teams = set(data['AwayTeamName'].unique())
    all_teams = sorted(home_teams.union(away_teams))
    
    # Get selected team from query parameter or user preferences
    team_name = request.args.get('team') or session.get('preferred_team')
    team_stats_detail = None
    
    if team_name:
        team_stats_detail = get_team_detail_stats(data, team_name)
    
    return render_template('team_detail.html',
                         all_teams=all_teams,
                         team_name=team_name,
                         team_stats=team_stats_detail,
                         divisions=divisions,
                         data_source_info=data_source_info)

@app.route('/player-stats', methods=['GET', 'POST'])
@user_required
def player_stats():
    """Dedicated player statistics page"""
    if data.empty:
        return render_template('player_stats.html', error="No data available", data_source_info=data_source_info)
    
    # Get selected division and team from form
    selected_division = request.form.get('division')
    selected_team = request.form.get('team')
    
    # Filter data based on division selection
    filtered_data = filter_data_by_division(data, selected_division)
    
    # Get comprehensive player statistics (filtered by team if selected)
    top_scorers = get_top_scorers(filtered_data, 50, division=None, team=selected_team)  # Get top 50 for comprehensive view
    highest_single_scores = get_highest_single_game_score(filtered_data, 10, division=None, team=selected_team)  # Now returns top 10
    top_three_pointers = get_top_three_pointers(filtered_data, 20, division=None, team=selected_team)
    top_foulers = get_top_foulers(filtered_data, 20, division=None, team=selected_team)
    
    # New basketball-specific statistics
    shooting_efficiency = get_player_shooting_efficiency(filtered_data, 20, division=None, team=selected_team)
    starter_bench_stats = get_starting_five_vs_bench_stats(filtered_data, division=None, team=selected_team)
    double_digit_scorers = get_double_digit_scorers(filtered_data, division=None, team=selected_team)
    consistent_scorers = get_consistent_scorers(filtered_data, division=None, team=selected_team)
    
    return render_template('player_stats.html',
                         top_scorers=top_scorers,
                         highest_single_scores=highest_single_scores,  # Updated variable name
                         top_three_pointers=top_three_pointers,
                         top_foulers=top_foulers,
                         shooting_efficiency=shooting_efficiency,
                         starter_bench_stats=starter_bench_stats,
                         double_digit_scorers=double_digit_scorers,
                         consistent_scorers=consistent_scorers,
                         divisions=divisions,
                         teams=teams,
                         selected_division=selected_division,
                         selected_team=selected_team,
                         data_source_info=data_source_info)

@app.route('/player-detail')
@user_required
def player_detail():
    """Individual player detail page with search and comprehensive statistics"""
    if data.empty:
        return render_template('player_detail.html', error="No data available", data_source_info=data_source_info)
    
    # Get all players for autocomplete
    all_players = get_all_players_list(data)
    
    # Get selected player from query parameter or user preferences
    player_name = request.args.get('player') or session.get('preferred_player')
    player_stats_detail = None
    
    if player_name:
        player_stats_detail = get_player_detail_stats(data, player_name)
    
    return render_template('player_detail.html',
                         all_players=all_players,
                         player_name=player_name,
                         player_stats=player_stats_detail,
                         divisions=divisions,
                         data_source_info=data_source_info)

@app.route('/referee-stats')
@user_required
def referee_stats():
    """Dedicated referee statistics page"""
    if data.empty:
        return render_template('referee_stats.html', error="No data available", data_source_info=data_source_info)
    
    # Get comprehensive referee statistics
    referee_stats_data = get_referee_statistics(data)
    referee_fouls = get_referee_fouls_per_game(data)
    referee_least_fouls = get_referees_least_fouls_per_game(data)
    referee_impact = get_referee_game_impact_analysis(data)
    
    return render_template('referee_stats.html',
                         referee_stats=referee_stats_data,
                         referee_fouls=referee_fouls,
                         referee_least_fouls=referee_least_fouls,
                         referee_impact=referee_impact,
                         divisions=divisions,
                         data_source_info=data_source_info)

@app.route('/referee-detail')
@user_required
def referee_detail():
    """Individual referee detail page with search and comprehensive statistics"""
    if data.empty:
        return render_template('referee_detail.html', error="No data available", data_source_info=data_source_info)
    
    # Get all referees for autocomplete
    all_referees = get_all_referees_list(data)
    
    # Get selected referee from query parameter
    referee_name = request.args.get('referee')
    referee_stats_detail = None
    
    if referee_name:
        referee_stats_detail = get_referee_detail_stats(data, referee_name)
    
    return render_template('referee_detail.html',
                         all_referees=all_referees,
                         referee_name=referee_name,
                         referee_stats=referee_stats_detail,
                         divisions=divisions,
                         data_source_info=data_source_info)

@app.route('/referee-performance-index')
@user_required
def referee_performance_index():
    """Referee Performance Index (RPI) page with comprehensive rankings"""
    if data.empty:
        return render_template('referee_performance_index.html', error="No data available", data_source_info=data_source_info)
    
    # Calculate Referee Performance Index
    rpi_data = calculate_referee_performance_index(data)
    
    # Prepare data for visualizations
    scatter_data = []
    radar_data = []
    
    if not rpi_data.empty:
        # Scatter plot data: Fairness vs Consistency
        for _, ref in rpi_data.iterrows():
            scatter_data.append({
                'name': ref['RefereeName'],
                'fairness': ref['FairnessScore'],
                'consistency': ref['ConsistencyScore'],
                'rpi': ref['RPI'],
                'games': ref['GamesRefereed']
            })
        
        # Radar data for top 10 referees
        for _, ref in rpi_data.head(10).iterrows():
            radar_data.append({
                'name': ref['RefereeName'],
                'fairness': ref['FairnessScore'],
                'consistency': ref['ConsistencyScore'],
                'control': ref['GameControlScore'],
                'experience': ref['ExperienceScore']
            })
    
    return render_template('referee_performance_index.html',
                         rpi_data=rpi_data,
                         scatter_data=scatter_data,
                         radar_data=radar_data,
                         divisions=divisions,
                         data_source_info=data_source_info)

@app.route('/deeper-analysis')
@user_required
def deeper_analysis():
    """Deep game analysis page with advanced metrics"""
    if data.empty:
        return render_template('deeper_analysis.html', error="No data available", divisions=divisions, data_source_info=data_source_info)
    
    # Get division filter from query parameters or user preferences
    division_filter = request.args.get('division') or session.get('preferred_division')
    
    # Filter data based on division selection
    filtered_data = filter_data_by_division(data, division_filter)
    
    # Get comprehensive deeper analysis with filtered data
    player_impact = get_player_game_impact_analysis(filtered_data, 20)
    foul_impact = get_player_foul_impact_analysis(filtered_data, 15) 
    player_combinations = get_best_player_combinations(filtered_data, 3)
    referee_impact = get_referee_game_impact_analysis(filtered_data)
    
    return render_template('deeper_analysis.html',
                         player_impact=player_impact,
                         foul_impact=foul_impact,
                         player_combinations=player_combinations,
                         referee_impact=referee_impact,
                         divisions=divisions,
                         selected_division=division_filter,
                         data_source_info=data_source_info)

@app.route('/fixtures')
def fixtures():
    """Fixtures page with games displayed as a matrix table"""
    if data.empty:
        return render_template('fixtures.html', error="No data available", divisions=divisions, data_source_info=data_source_info)
    
    # Get division filter from query parameters or user preferences
    # Default to "M-Division 1:" if no filter is provided (first time visit) and no preference set
    DEFAULT_DIVISION = "M-Division 1:"
    division_param = request.args.get('division')
    preferred_division = session.get('preferred_division')
    
    if division_param is None:
        # No explicit param - use preference or default
        if preferred_division:
            division_filter = preferred_division
            selected_division_param = preferred_division
        else:
            # First time visit - default to "M-Division 1:"
            division_filter = DEFAULT_DIVISION
            selected_division_param = DEFAULT_DIVISION
    elif division_param == "ALL":
        # User explicitly selected "All Divisions" - show all
        division_filter = None
        selected_division_param = "ALL"
    else:
        # Specific division selected
        division_filter = division_param
        selected_division_param = division_param
    
    # Get matrix data for fixtures
    matrix_data = get_fixtures_matrix_data(data, division_filter)
    
    # Also get traditional table data with filter applied
    fixtures_data = get_all_fixtures_data(data, division_filter)
    
    # Sort by date (oldest first)
    if not fixtures_data.empty and 'DateTime' in fixtures_data.columns:
        fixtures_data = fixtures_data.sort_values('DateTime', ascending=True)
    
    # Get closest games for each team
    closest_games = get_closest_games_by_team(data, division_filter)
    
    return render_template('fixtures.html',
                         fixtures=fixtures_data,
                         matrix_data=matrix_data,
                         divisions=divisions,
                         selected_division=selected_division_param,
                         data_source_info=data_source_info,
                         closest_games=closest_games)

@app.route('/game-detail/<game_id>')
@user_required
def game_detail(game_id):
    """Game detail page showing comprehensive information about a specific game"""
    if data.empty:
        return render_template('game_detail.html', error="No data available", data_source_info=data_source_info)
    
    # Get all games for autocomplete
    all_games = get_all_games_list(data)
    
    # Get game details
    game_details = get_game_details(data, game_id)
    
    if not game_details:
        return render_template('game_detail.html', error=f"Game {game_id} not found", all_games=all_games, game_id=game_id, data_source_info=data_source_info)
    
    # Generate funny game review
    game_review = generate_game_review(game_details)
    
    return render_template('game_detail.html',
                         game=game_details,
                         all_games=all_games,
                         game_id=game_id,
                         game_review=game_review,
                         data_source_info=data_source_info)

@app.route('/game-details')
@user_required
def game_details_search():
    """Game details search page with searchable functionality similar to player detail"""
    if data.empty:
        return render_template('game_details.html', error="No data available", data_source_info=data_source_info)
    
    # Get all games for autocomplete
    all_games = get_all_games_list(data)
    
    # Get selected game from query parameter
    game_id = request.args.get('game')
    game_details_data = None
    
    if game_id:
        game_details_data = get_game_details(data, game_id)
    
    return render_template('game_details.html',
                         all_games=all_games,
                         game_id=game_id,
                         game=game_details_data,
                         divisions=divisions,
                         data_source_info=data_source_info)

@app.route('/game-details/export/<game_id>')
@user_required
def export_game_details(game_id):
    """Export game details to Excel file"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from flask import send_file
        import io
        
        # Get game details
        game_details = get_game_details(data, game_id)
        
        if not game_details:
            return "Game not found", 404
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Game Statistics"
        
        # Game info header
        basic_info = game_details.get('basic_info', {})
        ws['A1'] = f"Game ID: {basic_info.get('game_id', '')}"
        ws['A2'] = f"{basic_info.get('home_team', '')} vs {basic_info.get('away_team', '')}"
        ws['A3'] = f"Final Score: {basic_info.get('final_score', '')}"
        ws['A4'] = f"Date: {basic_info.get('date_time', '')}"
        ws['A5'] = f"Division: {basic_info.get('division', '')}"
        
        # Style header
        for row in range(1, 6):
            ws[f'A{row}'].font = Font(bold=True, size=12)
        
        # Home team roster
        row = 7
        ws[f'A{row}'] = f"{basic_info.get('home_team', 'Home Team')} - Score: {basic_info.get('home_score', 0)}"
        ws[f'A{row}'].font = Font(bold=True, size=14, color="2980b9")
        ws[f'A{row}'].fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        
        row += 1
        # Header row
        headers = ['Player', 'Points', 'P', 'P1', 'P2', 'P3', 'T1', 'U1', 'U2', 'U3', 'GD', 'Total Fouls']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center')
        
        # Home team players
        home_roster = game_details.get('home_team_roster', [])
        for player in home_roster:
            row += 1
            player_name = player.get('player_name', '')
            if player.get('starting_five'):
                player_name += ' ⭐'
            
            values = [
                player_name,
                player.get('total_points', 0),
                '-' if player.get('p_fouls') is None else player.get('p_fouls', 0),
                '-' if player.get('p1_fouls') is None else player.get('p1_fouls', 0),
                '-' if player.get('p2_fouls') is None else player.get('p2_fouls', 0),
                '-' if player.get('p3_fouls') is None else player.get('p3_fouls', 0),
                '-' if player.get('t1_fouls') is None else player.get('t1_fouls', 0),
                '-' if player.get('u1_fouls') is None else player.get('u1_fouls', 0),
                '-' if player.get('u2_fouls') is None else player.get('u2_fouls', 0),
                '-' if player.get('u3_fouls') is None else player.get('u3_fouls', 0),
                '-' if player.get('gd_fouls') is None else player.get('gd_fouls', 0),
                player.get('total_fouls', 0)
            ]
            
            for col, value in enumerate(values, start=1):
                ws.cell(row=row, column=col, value=value)
        
        # Away team roster
        row += 3
        ws[f'A{row}'] = f"{basic_info.get('away_team', 'Away Team')} - Score: {basic_info.get('away_score', 0)}"
        ws[f'A{row}'].font = Font(bold=True, size=14, color="e67e22")
        ws[f'A{row}'].fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
        
        row += 1
        # Header row
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center')
        
        # Away team players
        away_roster = game_details.get('away_team_roster', [])
        for player in away_roster:
            row += 1
            player_name = player.get('player_name', '')
            if player.get('starting_five'):
                player_name += ' ⭐'
            
            values = [
                player_name,
                player.get('total_points', 0),
                '-' if player.get('p_fouls') is None else player.get('p_fouls', 0),
                '-' if player.get('p1_fouls') is None else player.get('p1_fouls', 0),
                '-' if player.get('p2_fouls') is None else player.get('p2_fouls', 0),
                '-' if player.get('p3_fouls') is None else player.get('p3_fouls', 0),
                '-' if player.get('t1_fouls') is None else player.get('t1_fouls', 0),
                '-' if player.get('u1_fouls') is None else player.get('u1_fouls', 0),
                '-' if player.get('u2_fouls') is None else player.get('u2_fouls', 0),
                '-' if player.get('u3_fouls') is None else player.get('u3_fouls', 0),
                '-' if player.get('gd_fouls') is None else player.get('gd_fouls', 0),
                player.get('total_fouls', 0)
            ]
            
            for col, value in enumerate(values, start=1):
                ws.cell(row=row, column=col, value=value)
        
        # Add legend
        row += 3
        ws[f'A{row}'] = "Foul Types Legend:"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        ws[f'A{row}'] = "P=Personal, P1=Shooting Foul, P2=Flagrant 1, P3=Flagrant 2"
        row += 1
        ws[f'A{row}'] = "T1=Technical, U1/U2/U3=Unsportsmanlike, GD=Game Disqualification"
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
            ws.column_dimensions[col].width = 12
        
        # Save to BytesIO
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Create filename
        home_team = basic_info.get('home_team', 'Home').replace(' ', '_')
        away_team = basic_info.get('away_team', 'Away').replace(' ', '_')
        filename = f"Game_{game_id}_{home_team}_vs_{away_team}.xlsx"
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logger.error(f"Error exporting game details: {e}")
        return f"Error exporting game: {str(e)}", 500


@app.route('/preferences', methods=['GET', 'POST'])
@user_required
def preferences():
    """User preferences page for setting default filters"""
    if request.method == 'POST':
        # Save preferences to session
        division = request.form.get('division') or None
        team = request.form.get('team') or None
        player = request.form.get('player') or None
        session['preferred_division'] = division
        session['preferred_team'] = team
        session['preferred_player'] = player

        
        # Validate and save theme preference
        theme = request.form.get('theme', 'default')
        if theme in VALID_THEMES:
            session['preferred_theme'] = theme
        else:
            session['preferred_theme'] = 'default'
        
        # Save to database if user is logged in with database account
        username = session.get('username')
        if username:
            success, message = update_user_preferences(
                username=username,
                division_name=division,
                team_name=team
            )
            if not success:
                logger.warning(f"Failed to update preferences in database: {message}")
        
        # Always redirect to preferences page (don't use user-provided URL)
        return redirect(url_for('preferences'))
    
    # Get all unique teams for dropdown
    home_teams = set(data['HomeTeamName'].unique()) if not data.empty else set()
    away_teams = set(data['AwayTeamName'].unique()) if not data.empty else set()
    all_teams = sorted(home_teams.union(away_teams))
    
    # Get all players for dropdown
    all_players = get_all_players_list(data)
    
    # Get current preferences from session
    current_prefs = {
        'division': session.get('preferred_division'),
        'team': session.get('preferred_team'),
        'player': session.get('preferred_player'),
        'theme': session.get('preferred_theme', 'default')
    }
    
    return render_template('preferences.html',
                         divisions=divisions,
                         all_teams=all_teams,
                         all_players=all_players,
                         current_prefs=current_prefs,
                         data_source_info=data_source_info)

@app.route('/help')
def help_page():
    """Help page with website features guideline"""
    return render_template('help.html',
                         data_source_info=data_source_info)

@app.route('/login', methods=['GET', 'POST'])
@limiter.limit("5 per 15 minutes", methods=["POST"])
def login():
    """Unified login page and authentication for all user levels"""
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        
        # Get IP address and user agent for logging
        ip_address = request.remote_addr
        user_agent = request.headers.get('User-Agent', '')
        
        # Try database authentication with logging info
        auth_success, user_data = authenticate_user(username, password, ip_address, user_agent)
        
        if auth_success:
            # Database authentication successful
            session['user_level'] = user_data.get('user_level', 'user')
            session['username'] = user_data['username']
            session.permanent = True  # Make session persistent
            
            # Load user preferences into session
            if user_data.get('division_name'):
                session['preferred_division'] = user_data['division_name']
            if user_data.get('team_name'):
                session['preferred_team'] = user_data['team_name']
            
            # Redirect to next URL if provided, otherwise to index (or admin for admins)
            # Check URL parameters first (from GET redirect), then form data (from POST with hidden field)
            # Use 'or' to handle None or empty string from either source
            next_url = request.args.get('next') or request.form.get('next') or None
            if next_url and next_url.startswith('/'):
                return redirect(next_url)
            elif user_data.get('user_level') == 'admin':
                return redirect(url_for('admin'))
            else:
                return redirect(url_for('index'))
        
        # Authentication failed
        # Preserve the next parameter in case of failed login
        next_url = request.args.get('next') or request.form.get('next', '')
        return render_template('login.html', 
                             error='Invalid username or password. Please try again.',
                             next=next_url)
    
    # GET request - show login form
    # If already authenticated, redirect appropriately
    if is_user_authenticated():
        if is_admin_authenticated():
            return redirect(url_for('admin'))
        return redirect(url_for('index'))
    
    # Pass the next parameter to the template so it can be preserved in the form
    next_url = request.args.get('next', '')
    return render_template('login.html', next=next_url)

# Keep old routes for backward compatibility (redirect to new unified login)
@app.route('/user/login', methods=['GET', 'POST'])
def user_login():
    """Redirect to unified login page"""
    return redirect(url_for('login', next=request.args.get('next')))

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    """Redirect to unified login page"""
    return redirect(url_for('login', next=request.args.get('next')))

@app.route('/logout')
def logout():
    """Unified logout for all user levels"""
    session.pop('user_level', None)
    session.pop('username', None)
    return redirect(url_for('index'))

# Keep old logout routes for backward compatibility
@app.route('/user/logout')
def user_logout():
    """Redirect to unified logout"""
    return redirect(url_for('logout'))

@app.route('/admin/logout')
def admin_logout():
    """Redirect to unified logout"""
    return redirect(url_for('logout'))

@app.route('/admin')
@admin_required
def admin():
    """Administration page with data statistics"""
    import os
    from datetime import datetime
    
    # Get season information
    season_info = get_season_info()
    website_config = get_website_config()
    
    # Calculate data statistics
    data_stats = {}
    
    if not data.empty:
        # Basic data statistics
        data_stats['total_games'] = len(data)
        data_stats['total_teams'] = len(set(list(data['HomeTeamName'].unique()) + list(data['AwayTeamName'].unique())))
        data_stats['total_divisions'] = len(data['GameDivisionDisplay'].unique())
        data_stats['total_data_points'] = len(data) * len(data.columns)
        data_stats['data_columns'] = len(data.columns)
        
        # Player statistics
        if 'PlayerName' in data.columns:
            data_stats['total_players'] = len(data['PlayerName'].unique())
        else:
            data_stats['total_players'] = 'N/A'
            
        # Date range
        if 'GameDate' in data.columns:
            try:
                data_stats['date_range_start'] = data['GameDate'].min()
                data_stats['date_range_end'] = data['GameDate'].max()
            except:
                data_stats['date_range_start'] = 'N/A'
                data_stats['date_range_end'] = 'N/A'
        else:
            data_stats['date_range_start'] = 'N/A'
            data_stats['date_range_end'] = 'N/A'
            
        # Score statistics
        if 'FinalHomeScore' in data.columns and 'FinalAwayScore' in data.columns:
            data_stats['avg_home_score'] = data['FinalHomeScore'].mean()
            data_stats['avg_away_score'] = data['FinalAwayScore'].mean()
            data_stats['highest_scoring_game'] = (data['FinalHomeScore'] + data['FinalAwayScore']).max()
        else:
            data_stats['avg_home_score'] = 'N/A'
            data_stats['avg_away_score'] = 'N/A'
            data_stats['highest_scoring_game'] = 'N/A'
            
        # Division statistics
        division_games = data.groupby('GameDivisionDisplay').size().to_dict()
        data_stats['division_games'] = division_games
        
    else:
        data_stats = {
            'total_games': 0,
            'total_teams': 0,
            'total_divisions': 0,
            'total_players': 0,
            'total_data_points': 0,
            'data_columns': 0,
            'date_range_start': 'N/A',
            'date_range_end': 'N/A',
            'avg_home_score': 'N/A',
            'avg_away_score': 'N/A',
            'highest_scoring_game': 'N/A',
            'division_games': {}
        }
    
    # File system statistics
    try:
        file_stats = {}
        # Check full-game-stats.csv using the constant from utils.py
        if os.path.exists(CSV_FILEPATH):
            mod_time = os.path.getmtime(CSV_FILEPATH)
            file_stats['full-game-stats.csv'] = {
                'size': os.path.getsize(CSV_FILEPATH),
                'modified': datetime.fromtimestamp(mod_time).strftime('%Y-%m-%d %H:%M:%S')
            }
        else:
            file_stats['full-game-stats.csv'] = None
        
        # Check file.csv in current directory (legacy check)
        file_csv_path = os.path.join(os.getcwd(), 'file.csv')
        if os.path.exists(file_csv_path):
            mod_time = os.path.getmtime(file_csv_path)
            file_stats['file.csv'] = {
                'size': os.path.getsize(file_csv_path),
                'modified': datetime.fromtimestamp(mod_time).strftime('%Y-%m-%d %H:%M:%S')
            }
        else:
            file_stats['file.csv'] = None
    except:
        file_stats = {}
    
    # List available season archives
    available_archives = list_available_archives()
    
    # Get user login statistics
    login_stats = get_login_statistics()
    users_with_login = get_users_with_login_info()
    recent_logins = get_recent_login_logs(limit=20)
    
    return render_template('admin.html',
                         data_stats=data_stats,
                         file_stats=file_stats,
                         data_source_info=data_source_info,
                         divisions=divisions,
                         season_info=season_info,
                         website_config=website_config,
                         available_archives=available_archives,
                         login_stats=login_stats,
                         users_with_login=users_with_login,
                         recent_logins=recent_logins)

@app.route('/admin/import-season', methods=['POST'])
@admin_required
def import_season_data():
    """Handle season archive import"""
    import os
    import tempfile
    
    if 'archive_file' not in request.files:
        return {'success': False, 'error': 'No file provided'}, 400
    
    file = request.files['archive_file']
    if file.filename == '':
        return {'success': False, 'error': 'No file selected'}, 400
    
    if not file.filename.endswith('.zip'):
        return {'success': False, 'error': 'Please upload a ZIP file'}, 400
    
    try:
        # Save uploaded file temporarily
        with tempfile.NamedTemporaryFile(delete=False, suffix='.zip') as tmp_file:
            file.save(tmp_file.name)
            
            # Import the archive
            result = import_season_archive(tmp_file.name)
            
            # Clean up temporary file
            os.unlink(tmp_file.name)
            
            if result['success']:
                return {
                    'success': True,
                    'message': f"Successfully imported {len(result['imported_files'])} files",
                    'season_id': result['season_id'],
                    'target_directory': result['target_directory']
                }
            else:
                return {'success': False, 'error': '; '.join(result['errors'])}, 400
                
    except Exception as e:
        return {'success': False, 'error': f'Import failed: {str(e)}'}, 500

@app.route('/admin/export-season', methods=['POST'])
@admin_required
def export_season_data():
    """Handle season data export"""
    import tempfile
    from flask import send_file
    
    try:
        # Get options from request
        include_raw = request.form.get('include_raw', 'false').lower() == 'true'
        
        # Create temporary file for export
        with tempfile.NamedTemporaryFile(delete=False, suffix='.zip') as tmp_file:
            tmp_path = tmp_file.name
        
        # Export the data
        result = export_season_archive(output_path=tmp_path, include_raw=include_raw)
        
        if result['success']:
            # Generate a nice filename
            from datetime import datetime
            season_id = result.get('season_id', 'unknown')
            timestamp = datetime.now().strftime('%Y%m%d-%H%M%S')
            filename = f"flbb-data-{season_id}-{timestamp}.zip"
            
            # Send file to user
            return send_file(
                tmp_path,
                as_attachment=True,
                download_name=filename,
                mimetype='application/zip'
            )
        else:
            # Log the full error for debugging
            logger.error(f"Export failed: {'; '.join(result['errors'])}")
            return {'success': False, 'error': 'Export failed. Please check server logs for details.'}, 400
            
    except Exception as e:
        # Log the full exception for debugging
        logger.error(f"Export failed with exception: {str(e)}", exc_info=True)
        return {'success': False, 'error': 'Export failed due to server error. Please contact administrator.'}, 500

# User management routes
@app.route('/admin/users')
@admin_required
def admin_users():
    """User management page - list all users"""
    users = list_users()
    return render_template('admin_users.html', users=users)

@app.route('/admin/users/create', methods=['POST'])
@admin_required
@limiter.limit("10 per hour")
def admin_create_user():
    """Create a new user with default password"""
    username = request.form.get('username', '').strip()
    user_level = request.form.get('user_level', 'user').strip()
    division_name = request.form.get('division_name', '').strip() or None
    team_name = request.form.get('team_name', '').strip() or None
    
    if not username:
        return jsonify({'success': False, 'error': 'Username is required'}), 400
    
    # Validate user level
    if user_level not in ('guest', 'user', 'admin'):
        return jsonify({'success': False, 'error': 'Invalid user level'}), 400
    
    # Default password as specified in requirements
    default_password = "kurwa"
    
    # Create the user
    success, message = create_user(
        username=username,
        password=default_password,
        user_level=user_level,
        division_name=division_name,
        team_name=team_name
    )
    
    if success:
        return jsonify({
            'success': True,
            'message': f"User '{username}' created successfully with level '{user_level}'"
        })
    else:
        return jsonify({'success': False, 'error': message}), 400

@app.route('/admin/users/update-level', methods=['POST'])
@admin_required
@limiter.limit("10 per hour")
def admin_update_user_level():
    """Update user authorization level"""
    username = request.form.get('username', '').strip()
    user_level = request.form.get('user_level', '').strip()
    
    if not username:
        return jsonify({'success': False, 'error': 'Username is required'}), 400
    
    if not user_level:
        return jsonify({'success': False, 'error': 'User level is required'}), 400
    
    # Update the user level
    success, message = update_user_level(username, user_level)
    
    if success:
        return jsonify({
            'success': True,
            'message': message
        })
    else:
        return jsonify({'success': False, 'error': message}), 400

@app.route('/admin/users/reset-password', methods=['POST'])
@admin_required
@limiter.limit("10 per hour")
def admin_reset_password():
    """Reset user password to default"""
    username = request.form.get('username', '').strip()
    
    if not username:
        return jsonify({'success': False, 'error': 'Username is required'}), 400
    
    # Default password as specified in requirements
    default_password = "kurwa"
    
    # Update the password
    success, message = update_user_password(username, default_password)
    
    if success:
        return jsonify({
            'success': True,
            'message': f"Password reset successfully for user '{username}'"
        })
    else:
        return jsonify({'success': False, 'error': message}), 400

@app.route('/admin/users/delete', methods=['POST'])
@admin_required
@limiter.limit("10 per hour")
def admin_delete_user():
    """Delete a user"""
    username = request.form.get('username', '').strip()
    
    if not username:
        return jsonify({'success': False, 'error': 'Username is required'}), 400
    
    # Delete the user
    success, message = delete_user(username)
    
    if success:
        return jsonify({
            'success': True,
            'message': message
        })
    else:
        return jsonify({'success': False, 'error': message}), 400

@app.route('/admin/foul-weights', methods=['GET'])
@admin_required
def admin_foul_weights():
    """Display foul weights editor"""
    weights = get_foul_weights()
    return jsonify(weights)

@app.route('/admin/foul-weights/update', methods=['POST'])
@admin_required
@limiter.limit(FOUL_WEIGHTS_UPDATE_RATE_LIMIT)
def admin_update_foul_weights():
    """Update foul weights"""
    try:
        # Get the submitted weights from the form
        weights = {}
        foul_types = ['P', 'P1', 'P2', 'P3', 'T1', 'U1', 'U2', 'U3', 'GD']
        
        for foul_type in foul_types:
            weight_value = request.form.get(f'weight_{foul_type}', '')
            if not weight_value:
                return jsonify({
                    'success': False, 
                    'error': f'Weight for {foul_type} is required'
                }), 400
            
            try:
                weights[foul_type] = float(weight_value)
            except ValueError:
                return jsonify({
                    'success': False,
                    'error': f'Invalid weight value for {foul_type}: {weight_value}'
                }), 400
        
        # Update the weights
        success, message = update_foul_weights(weights)
        
        if success:
            return jsonify({
                'success': True,
                'message': message
            })
        else:
            return jsonify({'success': False, 'error': message}), 400
            
    except Exception as e:
        logger.error(f"Error updating foul weights: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/user/change-password', methods=['GET', 'POST'])
@user_required
def user_change_password():
    """Allow users to change their own password"""
    if request.method == 'POST':
        current_password = request.form.get('current_password', '')
        new_password = request.form.get('new_password', '')
        confirm_password = request.form.get('confirm_password', '')
        
        # Get current username from session
        username = session.get('username')
        if not username:
            return render_template('change_password.html', 
                                 error='Session expired. Please login again.')
        
        # Validate inputs
        if not current_password or not new_password or not confirm_password:
            return render_template('change_password.html',
                                 error='All fields are required')
        
        if new_password != confirm_password:
            return render_template('change_password.html',
                                 error='New passwords do not match')
        
        if len(new_password) < 5:
            return render_template('change_password.html',
                                 error='New password must be at least 5 characters')
        
        # Verify current password
        auth_success, _ = authenticate_user(username, current_password)
        if not auth_success:
            return render_template('change_password.html',
                                 error='Current password is incorrect')
        
        # Update password
        success, message = update_user_password(username, new_password)
        
        if success:
            return render_template('change_password.html',
                                 success='Password changed successfully')
        else:
            return render_template('change_password.html',
                                 error=f'Failed to change password: {message}')
    
    # GET request - show form
    return render_template('change_password.html')

# API endpoints for hover tooltips
@app.route('/api/hover/player/<player_name>')
def api_player_hover(player_name):
    """API endpoint to get player hover statistics"""
    if data.empty:
        return jsonify({'error': 'No data available'}), 404
    
    # URL decode the player name
    player_name = unquote(player_name)
    
    stats = get_player_hover_stats(data, player_name)
    if stats is None:
        return jsonify({'error': 'Player not found'}), 404
    
    return jsonify(stats)

@app.route('/api/hover/team/<team_name>')
def api_team_hover(team_name):
    """API endpoint to get team hover statistics"""
    if data.empty:
        return jsonify({'error': 'No data available'}), 404
    
    # URL decode the team name
    team_name = unquote(team_name)
    
    stats = get_team_hover_stats(data, team_name)
    if stats is None:
        return jsonify({'error': 'Team not found'}), 404
    
    return jsonify(stats)

@app.route('/api/hover/referee/<referee_name>')
def api_referee_hover(referee_name):
    """API endpoint to get referee hover statistics"""
    if data.empty:
        return jsonify({'error': 'No data available'}), 404
    
    # URL decode the referee name
    referee_name = unquote(referee_name)
    
    stats = get_referee_hover_stats(data, referee_name)
    if stats is None:
        return jsonify({'error': 'Referee not found'}), 404
    
    return jsonify(stats)

@app.route('/api/hover/game/<game_id>')
def api_game_hover(game_id):
    """API endpoint to get game hover statistics"""
    if data.empty:
        return jsonify({'error': 'No data available'}), 404
    
    stats = get_game_hover_stats(data, game_id)
    if stats is None:
        return jsonify({'error': 'Game not found'}), 404
    
    return jsonify(stats)

@app.route('/api/hover/division/<division_name>')
def api_division_hover(division_name):
    """API endpoint to get division hover statistics"""
    if data.empty:
        return jsonify({'error': 'No data available'}), 404
    
    # URL decode the division name
    division_name = unquote(division_name)
    
    stats = get_division_hover_stats(data, division_name)
    if stats is None:
        return jsonify({'error': 'Division not found'}), 404
    
    return jsonify(stats)

if __name__ == '__main__':
    app.run(debug=True, port=5001)
